"""
Phase 4 — Compare extracted addresses against the original Excel and write results.

For each row that has a website:
    • UK address found, all non-null scraped fields match Excel → address_status = "match"
    • UK address found, any field differs → address_status = "updated"
      (Excel address columns replaced)
    • Scrape/extraction failed, Add1+Town present → address_status = "original"
      (keeps Excel address as-is)
    • Non-UK detected → address_status = "skipped_non_uk"
    • Scrape/extraction failed, no Excel address → address_status = "skipped_no_address"

Output: {original_stem}_validated.xlsx
"""

import json
import re
import string
from pathlib import Path

_FENCE_RE = re.compile(r"^```(?:json)?\s*|\s*```$", re.DOTALL)

import pandas as pd
from scrape_websites import UK_POSTCODE_RE

EXTRACTED_JSON = Path("extracted_addresses.jsonl")
UK_ADDRESSES_JSON = Path("uk_addresses.json")
SKIPPED_JSON = Path("skipped_addresses.json")
SCRAPED_SOURCES_JSON = Path("scraped_sources.json")

ADDRESS_FIELDS = ["add1", "add2", "town", "county", "postcode"]

# ── jsonl checkpoint helpers ────────────────────────────────────────────────────


def _read_jsonl(path: Path) -> dict:
    result = {}
    if not path.exists():
        return result
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                entry = json.loads(line)
                result[entry["key"]] = entry["value"]
            except (json.JSONDecodeError, KeyError):
                continue  # skip corrupted line — crash safety
    return result


# ── normalisation ─────────────────────────────────────────────────────────────

_PUNCT_TABLE = str.maketrans("", "", string.punctuation)


def _norm(value) -> str:
    """Strip, lowercase, remove punctuation, collapse whitespace."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip().lower().translate(_PUNCT_TABLE)
    return re.sub(r"\s+", " ", text).strip()


# ── helpers ───────────────────────────────────────────────────────────────────


def _col_map(df: pd.DataFrame) -> dict[str, str]:
    """
    Map canonical field name → actual DataFrame column name (case-insensitive).
    Only includes fields that exist in the DataFrame.
    """
    lower_cols = {c.strip().lower(): c for c in df.columns}
    return {
        field: lower_cols[field]
        for field in ADDRESS_FIELDS
        if field in lower_cols
    }


def _unwrap_raw(address: dict) -> dict:
    """
    If the dict is {"raw": "```json...```"} (extraction stored fenced text),
    attempt to strip the fences and re-parse. Returns the original dict on
    any failure so callers can still fall through to the _has_address_data check.
    """
    if set(address.keys()) != {"raw"}:
        return address
    raw_text = address.get("raw", "")
    if not isinstance(raw_text, str):
        return address
    try:
        clean = _FENCE_RE.sub("", raw_text).strip()
        parsed = json.loads(clean)
        if isinstance(parsed, dict):
            return parsed
    except (json.JSONDecodeError, TypeError):
        pass
    return address


def _has_address_data(address: dict) -> bool:
    """True if at least one address field is non-null/non-empty."""
    return any(address.get(f) for f in ADDRESS_FIELDS)


def _compare_row(
    excel_row: pd.Series,
    scraped: dict,
    col_map: dict[str, str],
) -> tuple[str, dict | None]:
    """
    Compare non-null scraped fields against the Excel row.

    Returns:
        ("match", None) — every non-null scraped field matches
        ("updated", {col: value}) — at least one field differs; full replacement dict
    """
    for field, col in col_map.items():
        scraped_val = scraped.get(field)
        if scraped_val is None:
            continue  # absent in scraped result — skip this field
        if _norm(scraped_val) != _norm(excel_row.get(col)):
            # any mismatch → build full replacement and return early
            updates = {col_map[f]: scraped.get(f) for f in ADDRESS_FIELDS if f in col_map}
            return "updated", updates

    return "match", None


# ── audit helpers ─────────────────────────────────────────────────────────────


def _concat_address_from_row(row: pd.Series, col_map: dict[str, str]) -> str:
    parts = [
        str(row.get(col_map[f])).strip()
        for f in ADDRESS_FIELDS
        if f in col_map
        and row.get(col_map[f]) is not None
        and not (isinstance(row.get(col_map[f]), float) and pd.isna(row.get(col_map[f])))
        and str(row.get(col_map[f])).strip()
    ]
    return ", ".join(parts)


def _concat_address_from_scraped(scraped: dict) -> str:
    parts = [
        str(scraped[f]).strip()
        for f in ADDRESS_FIELDS
        if scraped.get(f) and str(scraped.get(f)).strip()
    ]
    return ", ".join(parts)


def _page_label(source_url: str | None, base_url: str) -> str:
    if not source_url:
        return ""
    suffix = source_url.rstrip("/").replace(base_url.rstrip("/"), "").strip("/")
    return "/" + suffix if suffix else "homepage"


def _summary_status(skip_reason: str) -> str:
    low = skip_reason.lower()
    if "non-uk" in low or "non_uk" in low:
        return "skipped - non-UK"
    if "scrape" in low or "dead" in low:
        return "skipped - dead URL"
    return "skipped - no address found"


def _is_non_uk_reason(reason: str) -> bool:
    low = reason.lower()
    return "non-uk" in low or "non_uk" in low


def _has_excel_address(row: pd.Series, col_map: dict[str, str]) -> bool:
    """True if Add1 and Town are both non-empty in the Excel row."""
    add1_col = col_map.get("add1")
    town_col = col_map.get("town")
    if not add1_col or not town_col:
        return False
    return bool(_norm(row.get(add1_col))) and bool(_norm(row.get(town_col)))


# ── UK address signals ────────────────────────────────────────────────────────

_UK_COUNTIES = frozenset(c.lower() for c in [
    # England
    "Greater London", "Surrey", "Kent", "Essex", "Hampshire",
    "Hertfordshire", "Berkshire", "Buckinghamshire", "Oxfordshire", "Sussex",
    "East Sussex", "West Sussex", "Suffolk", "Norfolk", "Cambridgeshire",
    "Bedfordshire", "Northamptonshire", "Leicestershire", "Nottinghamshire",
    "Derbyshire", "Lincolnshire", "Yorkshire", "North Yorkshire", "South Yorkshire",
    "West Yorkshire", "East Yorkshire", "Lancashire", "Cheshire", "Merseyside",
    "Greater Manchester", "Tyne and Wear", "Durham", "Northumberland",
    "Cumbria", "Staffordshire", "Warwickshire", "West Midlands", "Worcestershire",
    "Herefordshire", "Shropshire", "Gloucestershire", "Somerset", "Wiltshire",
    "Dorset", "Devon", "Cornwall", "Bristol", "Bath", "Rutland",
    # Scotland
    "Aberdeenshire", "Angus", "Argyll", "Ayrshire", "Borders",
    "Caithness", "Clackmannanshire", "Dumfries", "Dunbartonshire", "Edinburgh",
    "Falkirk", "Fife", "Glasgow", "Highland", "Inverclyde", "Lanarkshire",
    "Lothian", "Moray", "Orkney", "Perth", "Renfrewshire", "Stirling", "Sutherland",
    "West Lothian", "Midlothian", "East Lothian",
    # Wales
    "Cardiff", "Swansea", "Newport", "Powys", "Ceredigion", "Pembrokeshire",
    "Carmarthenshire", "Glamorgan", "Gwynedd", "Conwy", "Denbighshire", "Flintshire",
    "Wrexham", "Monmouthshire", "Blaenau Gwent", "Caerphilly", "Merthyr Tydfil",
    "Neath", "Bridgend", "Vale of Glamorgan", "Rhondda", "Torfaen", "Isle of Anglesey",
    # Northern Ireland
    "Antrim", "Armagh", "Down", "Fermanagh", "Londonderry", "Tyrone",
    "Belfast", "Derry",
])

_UK_CITIES = frozenset(c.lower() for c in [
    "London", "Manchester", "Birmingham", "Leeds", "Glasgow", "Edinburgh", "Liverpool",
    "Bristol", "Sheffield", "Bradford", "Cardiff", "Coventry", "Nottingham", "Leicester",
    "Newcastle", "Brighton", "Portsmouth", "Southampton", "Reading", "Oxford", "Cambridge",
    "Aberdeen", "Dundee", "Inverness", "Belfast", "Derby", "Plymouth", "Exeter", "Norwich",
    "York", "Sunderland", "Wolverhampton", "Stoke", "Swansea", "Newport", "Luton",
    "Milton Keynes", "Northampton", "Peterborough", "Cheltenham", "Gloucester",
    "Worcester", "Hereford", "Shrewsbury", "Chester", "Preston", "Blackpool",
    "Blackburn", "Bolton", "Wigan", "Oldham", "Rochdale", "Stockport", "Salford",
    "Huddersfield", "Halifax", "Wakefield", "Barnsley", "Rotherham", "Doncaster",
    "Middlesbrough", "Durham", "Carlisle", "Ipswich", "Colchester",
    "Southend", "Basildon", "Watford", "St Albans", "Stevenage", "Hatfield",
    "Guildford", "Woking", "Crawley", "Hastings", "Eastbourne", "Worthing",
    "Bournemouth", "Poole", "Salisbury", "Bath", "Taunton", "Truro",
    "Newbury", "Basingstoke", "Winchester", "Chichester", "Maidstone", "Canterbury",
    "Tunbridge Wells", "Medway",
])


def is_uk_address(row: pd.Series, col_map: dict[str, str]) -> bool:
    """
    Return True if any positive UK signal is found in the row's address fields.

    Checks all address fields combined for:
        1. UK postcode regex match
        2. Known UK county/region substring
        3. Known UK city/town substring
    """
    parts = []
    for f in ADDRESS_FIELDS:
        if f not in col_map:
            continue
        val = row.get(col_map[f])
        try:
            s = "" if pd.isna(val) else str(val).strip()
        except (TypeError, ValueError):
            s = str(val).strip() if val is not None else ""
        parts.append(s)

    combined = " ".join(parts).lower()
    if not combined.strip():
        return False

    if UK_POSTCODE_RE.search(combined):
        return True
    if any(county in combined for county in _UK_COUNTIES):
        return True
    if any(city in combined for city in _UK_CITIES):
        return True

    return False


def _write_summary_sheet(
    output_path: Path,
    audit_rows: list[dict],
    stats: dict[str, int],
) -> None:
    """Append an 'Address Summary' tab to the existing workbook using openpyxl."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    wb = load_workbook(output_path)
    if "Address Summary" in wb.sheetnames:
        del wb["Address Summary"]
    ws = wb.create_sheet("Address Summary")

    no_address_count = sum(
        1 for r in audit_rows if r["status"] == "skipped - no address found"
    )
    non_uk_dead_count = sum(
        1 for r in audit_rows if r["status"] in ("skipped - non-UK", "skipped - dead URL")
    )
    total_processed = (
        stats["match"] + stats["updated"]
        + stats.get("original", 0)
        + stats.get("skipped_non_uk", 0)
        + stats.get("skipped_no_address", 0)
    )

    stat_rows = [
        ("Total URLs processed", total_processed),
        ("Addresses matched (website matches sheet)", stats["match"]),
        ("Addresses updated (website had different/better address)", stats["updated"]),
        ("Kept original Excel address (scrape found nothing)", stats.get("original", 0)),
        ("No address found on website (skipped)", no_address_count),
        ("Non-UK or dead URL (skipped)", non_uk_dead_count),
    ]
    for label, value in stat_rows:
        ws.append([label, value])
        ws.cell(ws.max_row, 1).font = Font(bold=True)

    ws.append([])  # blank separator

    headers = ["website", "status", "old_address", "new_address", "notes"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for row in audit_rows:
        ws.append([
            row["website"],
            row["status"],
            row["old_address"],
            row["new_address"],
            row["notes"],
        ])

    wb.save(output_path)
    print(f"Address Summary sheet written → {output_path}")


def _write_ready_skipped_sheets(output_path: Path, df: pd.DataFrame) -> None:
    """Append 'Ready' and 'Skipped' sheets to the existing workbook."""
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font

    _KEEP = frozenset({"match", "updated", "original"})
    _SKIP = frozenset({"skipped_non_uk", "skipped_no_address"})

    status_col = next(
        (c for c in df.columns if c.strip().lower() == "address_status"),
        None,
    )
    if status_col:
        ready_df = df[df[status_col].apply(
            lambda v: str(v).strip().lower() in _KEEP if pd.notna(v) else False
        )].reset_index(drop=True)
        skipped_df = df[df[status_col].apply(
            lambda v: str(v).strip().lower() in _SKIP if pd.notna(v) else False
        )].reset_index(drop=True)
    else:
        ready_df = df.copy()
        skipped_df = pd.DataFrame(columns=df.columns)

    ready_df = ready_df.fillna("").astype(str)
    skipped_df = skipped_df.fillna("").astype(str)
    wb = load_workbook(output_path)
    for sheet_name, frame in (("Ready", ready_df), ("Skipped", skipped_df)):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(frame, index=False, header=True)):
            ws.append(row)
            if r_idx == 0:
                for cell in ws[1]:
                    cell.font = Font(bold=True)

    wb.save(output_path)
    print(f"Ready/Skipped sheets written → {output_path}")


# ── main ──────────────────────────────────────────────────────────────────────


def run(
    input_file: str | Path,
    extracted: dict[str, dict] | None = None,
    uk_addresses: dict[str, dict] | None = None,
    skipped_reasons: dict[str, str] | None = None,
    skip_indices: set[str] | None = None,
) -> tuple[Path, dict[str, int]]:
    """
    Merge extracted/filtered addresses back into the original Excel.

    Parameters
    ----------
    input_file        Original .xlsx file.
    extracted         All extracted addresses (phase 2 output).
    uk_addresses      UK-only subset (phase 3 output).
    skipped_reasons   Mapping of row_idx → skip reason (phase 3 output).
    skip_indices      Row indices to leave untouched (already processed).

    Returns
    -------
    (output_path, stats)
        output_path — Path to the written *_validated.xlsx file.
        stats — {"match": N, "updated": N, "skipped": N, "no_website": N}
    """
    input_file = Path(input_file)
    df = pd.read_excel(input_file)

    # ── load intermediate data from disk when not passed in ───────────────
    if extracted is None:
        extracted = _read_jsonl(EXTRACTED_JSON)
    if uk_addresses is None:
        uk_addresses = json.loads(UK_ADDRESSES_JSON.read_text())
    if skipped_reasons is None:
        if SKIPPED_JSON.exists():
            skipped_reasons = json.loads(SKIPPED_JSON.read_text())
        else:
            # fallback: anything extracted but absent from uk_addresses was filtered
            skipped_reasons = {
                idx: "non-UK address"
                for idx in extracted
                if idx not in uk_addresses
            }

    col_map = _col_map(df)
    if not col_map:
        print(
            "Warning: none of the address columns "
            f"({', '.join(ADDRESS_FIELDS)}) found in the Excel — "
            "comparison will be a no-op."
        )

    website_col = next(
        (c for c in df.columns if c.strip().lower() == "website"),
        None,
    )

    scraped_sources: dict[str, str] = {}
    if SCRAPED_SOURCES_JSON.exists():
        scraped_sources = json.loads(SCRAPED_SOURCES_JSON.read_text())

    # ensure output columns exist
    if "address_status" not in df.columns:
        df["address_status"] = ""
    df["address_status"] = df["address_status"].astype(object)
    if "skip_reason" not in df.columns:
        df["skip_reason"] = pd.NA

    stats = {"match": 0, "updated": 0, "original": 0, "skipped_non_uk": 0, "skipped_no_address": 0, "no_website": 0}
    audit_rows: list[dict] = []

    for idx in df.index:
        row_key = str(idx)

        if skip_indices and row_key in skip_indices:
            continue

        # rows with no website were never scraped — leave untouched
        has_website = (
            website_col is not None
            and not pd.isna(df.at[idx, website_col])
            and str(df.at[idx, website_col]).strip()
        )
        if not has_website:
            stats["no_website"] += 1
            continue

        website = str(df.at[idx, website_col]).strip()
        old_address = _concat_address_from_row(df.loc[idx], col_map)

        if row_key in uk_addresses:
            addr = _unwrap_raw(uk_addresses[row_key])

            # guard against extraction failures stored as {"raw": ...}
            if not _has_address_data(addr):
                if _has_excel_address(df.loc[idx], col_map):
                    if is_uk_address(df.loc[idx], col_map):
                        df.at[idx, "address_status"] = "original"
                        stats["original"] += 1
                        audit_rows.append({
                            "website": website,
                            "status": "original - used excel address",
                            "old_address": old_address,
                            "new_address": "",
                            "notes": "extraction failed; kept existing address",
                        })
                    else:
                        df.at[idx, "address_status"] = "skipped_non_uk"
                        df.at[idx, "skip_reason"] = "no UK address signals found"
                        stats["skipped_non_uk"] += 1
                        audit_rows.append({
                            "website": website,
                            "status": "skipped - non-UK",
                            "old_address": old_address,
                            "new_address": "",
                            "notes": "no UK address signals found",
                        })
                else:
                    df.at[idx, "address_status"] = "skipped_no_address"
                    df.at[idx, "skip_reason"] = "extraction failed"
                    stats["skipped_no_address"] += 1
                    audit_rows.append({
                        "website": website,
                        "status": "skipped - no address found",
                        "old_address": old_address,
                        "new_address": "",
                        "notes": "extraction failed",
                    })
                continue

            status, updates = _compare_row(df.loc[idx], addr, col_map)
            source_url = scraped_sources.get(row_key)
            notes = _page_label(source_url, website) if source_url else ""

            df.at[idx, "address_status"] = status
            if updates:
                for col, val in updates.items():
                    df.at[idx, col] = val
            stats[status] += 1

            audit_rows.append({
                "website": website,
                "status": status,
                "old_address": old_address,
                "new_address": _concat_address_from_scraped(addr) if status == "updated" else "",
                "notes": notes,
            })

        else:
            reason = skipped_reasons.get(row_key, "scrape failed")
            if _is_non_uk_reason(reason):
                df.at[idx, "address_status"] = "skipped_non_uk"
                df.at[idx, "skip_reason"] = reason
                stats["skipped_non_uk"] += 1
                audit_rows.append({
                    "website": website,
                    "status": "skipped - non-UK",
                    "old_address": old_address,
                    "new_address": "",
                    "notes": reason,
                })
            elif _has_excel_address(df.loc[idx], col_map):
                if is_uk_address(df.loc[idx], col_map):
                    df.at[idx, "address_status"] = "original"
                    stats["original"] += 1
                    audit_rows.append({
                        "website": website,
                        "status": "original - used excel address",
                        "old_address": old_address,
                        "new_address": "",
                        "notes": reason,
                    })
                else:
                    df.at[idx, "address_status"] = "skipped_non_uk"
                    df.at[idx, "skip_reason"] = "no UK address signals found"
                    stats["skipped_non_uk"] += 1
                    audit_rows.append({
                        "website": website,
                        "status": "skipped - non-UK",
                        "old_address": old_address,
                        "new_address": "",
                        "notes": "no UK address signals found",
                    })
            else:
                df.at[idx, "address_status"] = "skipped_no_address"
                df.at[idx, "skip_reason"] = reason
                stats["skipped_no_address"] += 1
                audit_rows.append({
                    "website": website,
                    "status": "skipped - no address found",
                    "old_address": old_address,
                    "new_address": "",
                    "notes": reason,
                })

    output_path = input_file.parent / f"{input_file.stem}_validated.xlsx"
    df = df.fillna("").astype(str)
    df.to_excel(output_path, index=False)
    _write_summary_sheet(output_path, audit_rows, stats)
    _write_ready_skipped_sheets(output_path, df)

    print(
        f"Compare: match={stats['match']} updated={stats['updated']} "
        f"original={stats['original']} skipped_non_uk={stats['skipped_non_uk']} "
        f"skipped_no_address={stats['skipped_no_address']} no_website={stats['no_website']}"
    )
    print(f"Output written → {output_path}")
    return output_path, stats


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print(f"Usage: python {Path(__file__).name} <input.xlsx>")
        sys.exit(1)
    path, _ = run(sys.argv[1])
